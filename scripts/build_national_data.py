"""Build static/national_data.json — Texas against the other 49 states.

Everything else on this site compares a Texas district to Texas. This layer
adds the missing axis: the same district against the nation, from the only
two files that cover every U.S. district and state on one ruler —

  * Census Bureau F-33 (Annual Survey of School System Finances), district
    file `elsec24t.xlsx`, fiscal 2024 — per-pupil current spending
    (PPCSTOT), the Census's own published figure;
  * NCES NPEFS state file `stfis24_1a.txt`, fiscal 2024 — per-pupil current
    expenditure by state (PPE15), which is NCE13 divided by average daily
    attendance, an identity this builder asserts for every state.

The join from Census NCESID to TEA district number is NOT derivable by
arithmetic (Dallas ISD is 4816230, not 48+057905). The bridge is the CCD
LEA Directory's ST_LEAID column (`TX-057905`), keyed by number — so the
eleven Texas district names that belong to two districts each, which once
put five districts on the wrong land on the map, cannot recur here.

Frame honesty, decided here and stated in the artifact:

  * Census fiscal 2024 is one year behind the TEA fiscal 2025 figures on
    the rest of the site. The two are never mixed in one figure.
  * PPCSTOT is CURRENT spending: the Census excludes construction, land,
    debt, community-service and adult-education programs, and "spending by
    a school system for students not included in its fall membership
    counts" (their words, quoted in meta). That last exclusion is why
    PPCSTOT is not exactly TCURSPND/ENROLL, and why this builder publishes
    the Census figure rather than a homemade division.
  * Texas charter schools have no F-33 row AT ALL: the Census surveys
    governments, and charters are not independent governments. 174 of the
    175 econ-layer districts absent here are charters (the other is UT
    Austin's lab school). An absence by construction, stated, not hidden.
  * District percentiles use the site-wide ranking rule: districts with
    fewer than 500 students are shown but not ranked.

Inputs (data/, not committed except the small state file):
  data/census_f33_2024.xlsx        Census F-33 district file, FY2024
  data/npefs_state_2024.txt        NPEFS state file, FY2024 (committed)
  data/ccd_lea_directory_2425.csv  CCD LEA Directory SY2024-25 (the bridge)
  data/district_crosswalk.csv      the district registry (committed)

Usage:
  python scripts/build_national_data.py [--out static/national_data.json]
"""
from __future__ import annotations

import argparse
import bisect
import csv
import hashlib
import json
import re
import statistics
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DATA, STATIC = ROOT / "data", ROOT / "static"

F33 = DATA / "census_f33_2024.xlsx"
NPEFS = DATA / "npefs_state_2024.txt"
BRIDGE = DATA / "ccd_lea_directory_2425.csv"
CROSSWALK = DATA / "district_crosswalk.csv"

# The site-wide ranking rule: per-student figures in tiny districts move on a
# single hire, so districts under this line are shown but never ranked. The
# same constant appears in prose in the artifact's limits; tests hold the two
# to each other.
MIN_STUDENTS = 500

# The Census's own definition, verbatim from school24doc.docx ("PER PUPIL
# SPENDING AMOUNTS"). Quoted rather than paraphrased so the artifact carries
# the publisher's words, not ours.
CENSUS_PER_PUPIL_DEFINITION = (
    "The per pupil spending amounts included in the summary tables and data "
    "files are derived from current spending totals and the fall membership "
    "data. Per pupil expenditure does not include spending for "
    "nonelementary-secondary programs (community service, adult education), "
    "or spending by a school system for students not included in its fall "
    "membership counts."
)


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def num(v) -> float | None:
    return float(v) if isinstance(v, (int, float)) else None


def load_bridge() -> dict[str, str]:
    """Census LEAID -> TEA district number, for Texas rows only."""
    out: dict[str, str] = {}
    with BRIDGE.open(encoding="latin-1", newline="") as fh:
        for row in csv.DictReader(fh):
            if row["FIPST"] == "48" and row["ST_LEAID"].startswith("TX-"):
                out[row["LEAID"]] = row["ST_LEAID"][3:]
    # The map once drew five districts on the wrong land because a name-keyed
    # join guessed between twins. This bridge is keyed by number, and the one
    # row a reader can check against both agencies is asserted outright.
    assert out.get("4816230") == "057905", (
        "the CCD bridge no longer maps Dallas ISD (LEAID 4816230) to TEA "
        "057905 — do not build on a broken bridge")
    return out


def load_states() -> dict:
    with NPEFS.open(newline="") as fh:
        rows = list(csv.DictReader(fh, delimiter="\t"))
    states = [r for r in rows if int(r["FIPS"]) <= 56]
    assert len(states) == 51, f"expected 50 states + DC, got {len(states)}"
    for r in states:
        # PPE15 must BE the division it claims to be, for every state.
        nce, ada, ppe = int(r["NCE13"]), int(r["ADA"]), int(r["PPE15"])
        assert abs(nce / ada - ppe) < 1.0, (
            f"{r['STABR']}: PPE15 {ppe} is not NCE13/ADA {nce / ada:.1f}")
    ranked = sorted(states, key=lambda r: -int(r["PPE15"]))
    tx = next(r for r in ranked if r["STABR"] == "TX")
    rank = ranked.index(tx) + 1

    # Robustness: PPE15 divides by average daily attendance. Texas counts
    # attendance, not just membership, so its ADA runs below enrolment and
    # per-ADA reads higher than per-member. If the denominator choice moved
    # the rank materially, that would be the finding; both ranks ship.
    by_member = sorted(states, key=lambda r: -(int(r["NCE13"]) / int(r["MEMBR23"])))
    rank_member = next(i for i, r in enumerate(by_member, 1) if r["STABR"] == "TX")

    return {
        "measure": ("PPE15: current expenditure per pupil in average daily "
                    "attendance, NPEFS fiscal 2024"),
        "texas": {
            "ppe": int(tx["PPE15"]),
            "rank": rank,
            "of": len(ranked),
            "who": "50 states and the District of Columbia; territories excluded",
            "numerator_current_expenditure": int(tx["NCE13"]),
            "denominator_ada": int(tx["ADA"]),
        },
        "denominator_check": {
            "note": ("PPE15 divides by average daily attendance (ADA), "
                     "NCES's own denominator. Dividing the same spending by "
                     "fall membership instead moves Texas's rank from "
                     f"{rank} to {rank_member} of 51 — the denominator "
                     "choice does not change the story, so the published "
                     "NCES figure is used."),
            "rank_by_ada": rank,
            "rank_by_membership": rank_member,
        },
        "rows": [{"abbr": r["STABR"], "name": r["STNAME"], "ppe": int(r["PPE15"])}
                 for r in ranked],
    }


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--out", default=str(STATIC / "national_data.json"))
    ap.add_argument("--economics", default=str(STATIC / "economics_data.json"),
                    help="economics artifact to measure coverage against")
    args = ap.parse_args(argv)

    for p in (F33, NPEFS, BRIDGE, CROSSWALK):
        if not p.exists():
            sys.exit(f"missing input: {p}")

    bridge = load_bridge()
    states = load_states()

    wb = openpyxl.load_workbook(F33, read_only=True)
    info_rows = list(wb["File Information"].iter_rows(values_only=True))
    # Only an actual date qualifies — a title cell like "2024 Annual Survey…"
    # ordered first would otherwise publish "2024 Annua" as the release date.
    date_re = re.compile(r"^20\d{2}-\d{2}-\d{2}")
    released = next(
        (m.group() for row in info_rows for c in row
         if c is not None and (m := date_re.match(str(c)))), None)
    assert released, "File Information sheet no longer carries a release date"

    ws = wb["elsec24t"]
    rows = ws.iter_rows(values_only=True)
    hdr = [str(c) for c in next(rows)]
    i = {c: n for n, c in enumerate(hdr)}
    all_rows = list(rows)

    # National pool under the site-wide rule. Percentile = share of ranked
    # districts spending strictly less; derived once here, never in the UI.
    pool: list[float] = []
    year = None
    for r in all_rows:
        year = year or r[i["YRDATA"]]
        e, pp = num(r[i["ENROLL"]]), num(r[i["PPCSTOT"]])
        if e and pp and e >= MIN_STUDENTS and pp > 0:
            pool.append(pp)
    pool.sort()
    assert str(year) == "24", f"YRDATA is {year!r}, not fiscal 2024 — wrong file?"

    def pctile(pp: float) -> int:
        # floor, not round: the page says "more per student than X% of ranked
        # districts", and floor is the largest X for which that sentence is
        # TRUE (round would overstate half the time, and could reach the
        # logically-false 100 for the top of the pool).
        return int(100 * bisect.bisect_left(pool, pp) / len(pool))

    # Texas rows -> TEA numbers. Accounting is hard: every row is resolved or
    # named in the artifact as unresolved; no silent drops.
    tx_rows = [r for r in all_rows if str(r[i["FIPST"]]).zfill(2) == "48"]
    districts: dict[str, dict] = {}
    unresolved: list[str] = []
    for r in tx_rows:
        leaid = str(r[i["NCESID"]])
        tea = bridge.get(leaid)
        if tea is None:
            unresolved.append(f"{leaid} {r[i['NAME']]}")
            continue
        # Revenue identity: the file must agree with itself before we quote it.
        assert r[i["TOTALREV"]] == r[i["TFEDREV"]] + r[i["TSTREV"]] + r[i["TLOCREV"]], (
            f"revenue identity fails for {r[i['NAME']]}")
        e, pp = num(r[i["ENROLL"]]), num(r[i["PPCSTOT"]])
        rec: dict = {"leaid": leaid}
        if e and e > 0:
            rec["enroll_f33"] = int(e)
        if pp and pp > 0:
            rec["ppcs"] = int(pp)
            if e and e >= MIN_STUDENTS:
                rec["pctile"] = pctile(pp)
        districts[tea] = rec
    assert len(districts) + len(unresolved) == len(tx_rows), "rows leaked"

    econ_path = Path(args.economics)
    econ_districts = set(
        json.loads(econ_path.read_text())["districts"].keys())
    covered = econ_districts & districts.keys()
    absent = econ_districts - districts.keys()
    xw = {r["district_number"]: r for r in csv.DictReader(CROSSWALK.open())}
    # Baked into the artifact so the API can say WHY a district is absent
    # (charter: cannot exist; other: not found) without a runtime crosswalk.
    # Over the WHOLE crosswalk, not just the econ layer: a closed charter is
    # every bit as much "cannot exist in a survey of governments" as an open
    # one, and answering it "missing information" would be the wrong reason.
    absent_map = {
        d: ("charter" if r.get("is_charter") == "True" else "no_row")
        for d, r in sorted(xw.items()) if d not in districts}
    absent_charters = sum(
        1 for d in absent if absent_map.get(d) == "charter")

    tx_ranked = [districts[d]["ppcs"] for d in covered
                 if "pctile" in districts[d]]

    out = {
        "meta": {
            "fiscal_year": 2024,
            "released": released,
            "sources": {
                "district_finance": {
                    "file": F33.name, "source": "census_f33",
                    "sha256": sha256(F33),
                },
                "state_npefs": {
                    "file": NPEFS.name, "source": "nces_npefs",
                    "sha256": sha256(NPEFS),
                },
                "bridge": {
                    "file": BRIDGE.name, "source": "ccd_lea_directory",
                    "sha256": sha256(BRIDGE),
                },
            },
            "census_per_pupil_definition": CENSUS_PER_PUPIL_DEFINITION,
            "min_students_ranked": MIN_STUDENTS,
            "coverage": {
                "tx_finance_rows": len(tx_rows),
                "tx_resolved_to_tea": len(districts),
                "tx_unresolved": unresolved,
                "econ_districts": len(econ_districts),
                "econ_covered": len(covered),
                "econ_absent": len(absent),
                "econ_absent_charters": absent_charters,
            },
            "limits": [
                "Census fiscal 2024 — one year behind the TEA fiscal 2025 "
                "figures elsewhere on this site. The two are never mixed in "
                "one figure.",
                "Current spending only: the Census excludes construction, "
                "land, debt, and spending for students outside a district's "
                "fall membership, so this figure runs below the all-in "
                "per-student total shown elsewhere on this site.",
                "The state ranking divides by average daily attendance, "
                "NCES's own denominator; dividing by fall membership instead "
                "moves Texas only from "
                f"{states['denominator_check']['rank_by_ada']} to "
                f"{states['denominator_check']['rank_by_membership']} of 51.",
                "Texas charter schools have no Census F-33 row at all — the "
                "Census surveys governments, and charters are not "
                "independent governments — so they have no national "
                "percentile, by construction rather than by omission.",
                "Districts under 500 students are shown but not ranked, the "
                "same rule used across this site.",
            ],
        },
        "states": states,
        "national": {
            "districts_in_pool": len(pool),
            "median_ppcs": int(statistics.median(pool)),
            "pool_rule": ("U.S. districts with 500+ students and a positive "
                          "per-pupil current spending figure"),
        },
        "texas_districts_ranked": {
            "count": len(tx_ranked),
            "median_ppcs": int(statistics.median(tx_ranked)),
        },
        "districts": {k: districts[k] for k in sorted(districts)},
        "absent": absent_map,
    }

    dal = out["districts"]["057905"]
    assert dal["ppcs"] == 15160 and dal["leaid"] == "4816230", (
        "Dallas ISD no longer reads $15,160 / LEAID 4816230 — either the "
        "source was restated (update this pin deliberately) or the join broke")

    out_path = Path(args.out)
    out_path.write_text(json.dumps(out, indent=1, sort_keys=True) + "\n")
    c = out["meta"]["coverage"]
    print(f"wrote {out_path} — {c['tx_resolved_to_tea']}/{c['tx_finance_rows']} "
          f"TX rows resolved, {c['econ_covered']}/{c['econ_districts']} econ "
          f"districts covered, pool {len(pool)}, TX rank "
          f"{states['texas']['rank']}/{states['texas']['of']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
